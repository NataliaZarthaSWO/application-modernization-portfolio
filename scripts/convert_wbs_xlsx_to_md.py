from __future__ import annotations

import argparse
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


HEADER_KEYWORDS = {
    "responsible",
    "stage",
    "service",
    "activity",
    "tool",
    "hours",
    "accelerator",
    "aws",
    "azure",
    "gcp",
    "effort",
}


def normalize_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    if not text:
        return ""
    lines = [" ".join(line.split()) for line in text.split("\n") if line.strip()]
    return "<br>".join(lines)


def escape_md(text: str) -> str:
    return text.replace("|", "\\|")


def trim_rows(rows: list[list[str]]) -> list[list[str]]:
    max_len = 0
    for row in rows:
        for index in range(len(row) - 1, -1, -1):
            if row[index]:
                max_len = max(max_len, index + 1)
                break
    return [row[:max_len] for row in rows]


def detect_header_index(rows: list[list[str]]) -> int | None:
    best_index = None
    best_score = -1

    for index, row in enumerate(rows[:10]):
        nonempty = [cell for cell in row if cell]
        if len(nonempty) < 2:
            continue

        score = 0
        for cell in nonempty:
            lowered = cell.lower()
            if any(keyword in lowered for keyword in HEADER_KEYWORDS):
                score += 3
            elif len(cell) <= 30:
                score += 1

        if score > best_score:
            best_index = index
            best_score = score

    return best_index


def build_headers(row: list[str]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}

    for index, cell in enumerate(row, start=1):
        header = cell or f"Column {index}"
        count = seen.get(header, 0)
        seen[header] = count + 1
        if count:
            header = f"{header} {count + 1}"
        headers.append(header)

    return headers


def active_columns(headers: list[str], data_rows: list[list[str]]) -> list[int]:
    keep: list[int] = []
    for index, header in enumerate(headers):
        if header and (any(index < len(row) and row[index] for row in data_rows) or header.startswith("Column ") is False):
            keep.append(index)
    return keep


def render_table(headers: list[str], rows: list[list[str]]) -> list[str]:
    lines = [
        "| " + " | ".join(escape_md(header) for header in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in rows:
        padded = row + [""] * (len(headers) - len(row))
        lines.append("| " + " | ".join(escape_md(value) for value in padded[: len(headers)]) + " |")
    return lines


def render_ai_powered_ama(rows: list[list[str]]) -> str:
    nonempty_rows = [row for row in trim_rows(rows) if any(cell for cell in row)]
    section = ["## Hoja: AI-Powered AMA", "", "### Resumen de esfuerzo", ""]

    summary_headers = [cell for cell in nonempty_rows[0] if cell]
    summary_rows = []
    for row in nonempty_rows[1:]:
        if row and row[0] == "#":
            break
        values = [cell for cell in row if cell]
        if len(values) == 1:
            section.append(f"_{values[0]}_")
            section.append("")
            continue
        if len(values) == 2:
            summary_rows.append(values)

    section.extend(render_table(summary_headers, summary_rows))
    section.extend(["", "### WBS", ""])

    header_index = next(index for index, row in enumerate(nonempty_rows) if row and row[0] == "#")
    headers = build_headers(nonempty_rows[header_index])
    data_rows = nonempty_rows[header_index + 1 :]
    keep = [index for index, header in enumerate(headers) if header]
    headers = [headers[index] for index in keep]
    data_rows = [[row[index] if index < len(row) else "" for index in keep] for row in data_rows]

    section.extend(render_table(headers, data_rows))
    return "\n".join(section)


def render_bullets(rows: list[list[str]]) -> list[str]:
    lines: list[str] = []
    for row in rows:
        values = [escape_md(cell) for cell in row if cell]
        if values:
            lines.append(f"- {' | '.join(values)}")
    return lines or ["- Sin contenido legible"]


def sheet_to_markdown(title: str, rows: list[list[str]]) -> str:
    if title == "AI-Powered AMA":
        return render_ai_powered_ama(rows)

    nonempty_rows = [row for row in trim_rows(rows) if any(cell for cell in row)]
    section = [f"## Hoja: {title}"]

    if not nonempty_rows:
        section.append("_Sin contenido legible._")
        return "\n".join(section)

    header_index = detect_header_index(nonempty_rows)
    if header_index is None:
        section.extend(render_bullets(nonempty_rows))
        return "\n".join(section)

    notes = nonempty_rows[:header_index]
    headers = build_headers(nonempty_rows[header_index])
    data_rows = [row for row in nonempty_rows[header_index + 1 :] if any(cell for cell in row)]

    if notes:
        section.append("### Notas")
        section.extend(render_bullets(notes))
        section.append("")

    if not data_rows:
        section.append("_Sin filas de datos debajo del encabezado._")
        return "\n".join(section)

    keep = active_columns(headers, data_rows)
    headers = [headers[index] for index in keep]
    data_rows = [[row[index] if index < len(row) else "" for index in keep] for row in data_rows]

    populated_per_row = [sum(1 for value in row if value) for row in data_rows]
    average_populated = mean(populated_per_row) if populated_per_row else 0

    if len(headers) > 15 or average_populated < 2:
        section.append("### Contenido")
        section.extend(render_bullets([headers] + data_rows))
        return "\n".join(section)

    section.extend(render_table(headers, data_rows))
    return "\n".join(section)


def workbook_to_markdown(source: Path, sheet_name: str | None = None) -> str:
    workbook = load_workbook(source, data_only=True)
    sections = [f"# {source.name}", "", "_Fuente: plantilla Excel convertida a Markdown_", ""]

    for worksheet in workbook.worksheets:
        if sheet_name and worksheet.title != sheet_name:
            continue
        rows = [[normalize_cell(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
        sections.append(sheet_to_markdown(worksheet.title, rows))
        sections.append("")

    return "\n".join(sections).rstrip() + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert a WBS Excel workbook to Markdown.")
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--sheet", dest="sheet_name", help="Convert only the named worksheet.")
    args = parser.parse_args()

    markdown = workbook_to_markdown(args.source, args.sheet_name)
    args.output.write_text(markdown, encoding="utf-8")


if __name__ == "__main__":
    main()